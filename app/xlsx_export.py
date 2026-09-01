"""
xlsx_export.py — construiește fișierul .xlsx cerut de pagina /export: toate
câmpurile din "Valori curente" (fișa stației) până la "Software instalat",
pentru setul de stații filtrat acolo (aceleași filtre ca la /statii).

openpyxl e singura dependență nouă (pură Python, fără compilare — vezi
requirements.txt și vendor_wheels/, unde e adăugată pentru instalarea offline
pe stația de domeniu, la fel ca Flask).
"""

import io
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from formatting import fmt_bool, fmt_dt  # noqa: E402

_HEADER_FONT = Font(bold=True)

STATION_COLUMNS = [
    ("name", "Stație"), ("ou_path", "OU"), ("ad_description", "Descriere AD"),
    ("status", "Status"), ("error_message", "Eroare"), ("level", "Nivel colectare"),
    ("manufacturer", "Producător"), ("model", "Model"), ("serial_number", "Serie"),
    ("bios_version", "BIOS"), ("cpu_name", "CPU"), ("ram_total_mb", "RAM (MB)"),
    ("os_caption", "OS"), ("os_build", "Build OS"), ("os_display_version", "Versiune afișată OS"),
    ("os_arch", "Arhitectură"), ("last_boot_fmt", "Ultima pornire"), ("uptime_days", "Uptime (zile)"),
    ("ip_address", "IP"), ("mac_address", "MAC"), ("dhcp_fmt", "DHCP"),
    ("logged_on_user_fmt", "Utilizator logat"), ("last_logged_on_user_fmt", "Ultimul utilizator"),
    ("av_name", "Antivirus"), ("av_enabled_fmt", "AV activ"), ("av_up_to_date_fmt", "AV la zi"),
    ("reboot_pending_fmt", "Reboot în așteptare"), ("wu_last_success_fmt", "Ultima actualizare Windows"),
    ("collected_at_fmt", "Colectat la"),
    ("c_free_pct", "Liber C: (%)"), ("c_free_mb", "Liber C: (MB)"), ("c_size_mb", "Total C: (MB)"),
]

DISK_COLUMNS = [
    ("host_name", "Stație"), ("device_id", "Unitate"), ("volume_name", "Volum"),
    ("size_mb", "Total (MB)"), ("free_mb", "Liber (MB)"), ("free_pct", "Liber (%)"),
]

AV_COLUMNS = [
    ("host_name", "Stație"), ("name", "Nume"), ("enabled_fmt", "Activ"),
    ("up_to_date_fmt", "La zi"), ("signature_date_fmt", "Dată semnătură"),
]

SOFTWARE_MACHINE_COLUMNS = [
    ("host_name", "Stație"), ("name", "Nume"), ("version", "Versiune"),
    ("publisher", "Producător"), ("install_date", "Instalat"), ("scope", "Scop"),
]

SOFTWARE_USER_COLUMNS = [
    ("host_name", "Stație"), ("user_name", "Utilizator"), ("name", "Nume"),
    ("version", "Versiune"), ("publisher", "Producător"), ("install_date", "Instalat"),
]


def _write_sheet(ws, columns, rows):
    ws.append([label for _, label in columns])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append([row.get(key) for key, _ in columns])
    for i, (_, label) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(label) + 2, 12), 40)
    ws.freeze_panes = "A2"


def _station_row(r):
    d = dict(r)
    d["last_boot_fmt"] = fmt_dt(r["last_boot"])
    d["dhcp_fmt"] = fmt_bool(r["dhcp_enabled"])
    d["logged_on_user_fmt"] = r["logged_on_user_display_name"] or r["logged_on_user"] or "—"
    d["last_logged_on_user_fmt"] = r["last_logged_on_user_display_name"] or r["last_logged_on_user"] or "—"
    has_av = bool(r["av_name"])
    d["av_enabled_fmt"] = fmt_bool(r["av_enabled"], "activ", "dezactivat") if has_av else "—"
    d["av_up_to_date_fmt"] = fmt_bool(r["av_up_to_date"], "la zi", "vechi") if has_av else "—"
    # reboot_pending/wu_last_success nu există decât la Nivel 2 (registry) —
    # la Nivel 1 rămân NULL în DB, indistinguibil de "nu"/necunoscut, deci
    # etichetăm explicit ca indisponibil (§5.7).
    is_l2 = r["level"] == 2
    d["reboot_pending_fmt"] = fmt_bool(r["reboot_pending"]) if is_l2 else "indisponibil la Nivel 1"
    d["wu_last_success_fmt"] = fmt_dt(r["wu_last_success"]) if is_l2 else "indisponibil la Nivel 1"
    d["collected_at_fmt"] = fmt_dt(r["collected_at"])
    return d


def _av_row(r):
    d = dict(r)
    d["enabled_fmt"] = fmt_bool(r["enabled"], "activ", "dezactivat")
    d["up_to_date_fmt"] = fmt_bool(r["up_to_date"], "la zi", "vechi")
    d["signature_date_fmt"] = fmt_dt(r["signature_date"])
    return d


def build_workbook(stations, disks, antivirus, software_machine, software_user) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valori curente"
    _write_sheet(ws, STATION_COLUMNS, [_station_row(r) for r in stations])

    _write_sheet(wb.create_sheet("Discuri"), DISK_COLUMNS, [dict(r) for r in disks])
    _write_sheet(wb.create_sheet("Antivirus"), AV_COLUMNS, [_av_row(r) for r in antivirus])
    _write_sheet(wb.create_sheet("Software (mașină)"), SOFTWARE_MACHINE_COLUMNS,
                 [dict(r) for r in software_machine])
    _write_sheet(wb.create_sheet("Software (utilizator)"), SOFTWARE_USER_COLUMNS,
                 [dict(r) for r in software_user])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
